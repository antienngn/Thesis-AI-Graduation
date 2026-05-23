"""build_trace_compare_xlsx.py — aggregate SERVE_OPTXXX vs SERVE_DUAL_TEST.

Output sheets:
  - all_runs          : long format, 1 row per (scheduler, qps)
  - nice_compare      : pivot ngang dual vs opt với ratio d/o + winner
  - {metric}          : pivot từng metric (qps × scheduler)

Metrics:
  duration, completed, request_throughput, output_throughput,
  median_ttft_ms, mean_ttft_ms, p99_ttft_ms,
  median_tpot_ms, mean_tpot_ms, p99_tpot_ms,
  mean_nlat_ms, median_nlat_ms
"""
import argparse
import glob
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SOURCES = {
    "dual1.0": "SERVE_DUAL_TEST",
    "opt-xxx": "SERVE_OPTXXX",
}
RATES = [2, 4, 8, 16, 32, 64]

METRIC_COLS = [
    "duration", "completed", "request_throughput", "output_throughput",
    "median_ttft_ms", "mean_ttft_ms", "p99_ttft_ms",
    "median_tpot_ms", "mean_tpot_ms", "p99_tpot_ms",
    "mean_nlat_ms", "median_nlat_ms",
]

PIVOT_SHEETS = [
    "median_ttft_ms", "mean_ttft_ms", "p99_ttft_ms",
    "median_tpot_ms", "mean_tpot_ms", "p99_tpot_ms",
    "mean_nlat_ms", "median_nlat_ms",
    "output_throughput", "request_throughput", "duration",
]


def compute_tpot_raw(d):
    """TPOT = mean(itl) per request, sau đó ms."""
    itls = d.get("itls") or []
    tpots = [sum(x) / len(x) for x in itls if len(x) > 0]
    if not tpots:
        return float("nan"), float("nan"), float("nan")
    a = np.asarray(tpots) * 1000.0
    return float(a.mean()), float(np.median(a)), float(np.percentile(a, 99))


def compute_nlat_raw(d):
    """N_lat = (TTFT + Σ ITL) / (len(itls)+1) per request, ms."""
    ttfts = d.get("ttfts") or []
    itls = d.get("itls") or []
    if not (ttfts and itls):
        return float("nan"), float("nan")
    vals = []
    for i in range(len(ttfts)):
        n_raw = len(itls[i]) + 1
        if n_raw > 0:
            vals.append((ttfts[i] + sum(itls[i])) / n_raw)
    if not vals:
        return float("nan"), float("nan")
    a = np.asarray(vals) * 1000.0
    return float(a.mean()), float(np.median(a))


def load_all(base: Path) -> pd.DataFrame:
    rows = []
    for sched, subdir in SOURCES.items():
        for r in RATES:
            pat = base / subdir / f"r{r}" / "*.json"
            files = sorted(glob.glob(str(pat)))
            if not files:
                continue
            with open(files[-1]) as fh:
                d = json.load(fh)
            mean_tpot, med_tpot, p99_tpot = compute_tpot_raw(d)
            mean_nlat, med_nlat = compute_nlat_raw(d)
            out_thp = (sum(len(x) + 1 for x in (d.get("itls") or []))
                       / d["duration"]) if d.get("duration") else None
            rows.append({
                "scheduler": sched,
                "qps": float(r),
                "source_dir": subdir,
                "file": Path(files[-1]).name,
                "model": d.get("model_id"),
                "num_prompts": d.get("num_prompts"),
                "completed": d.get("completed"),
                "duration": d.get("duration"),
                "request_throughput": d.get("request_throughput"),
                "output_throughput": out_thp,
                "median_ttft_ms": d.get("median_ttft_ms"),
                "mean_ttft_ms": d.get("mean_ttft_ms"),
                "p99_ttft_ms": d.get("p99_ttft_ms"),
                "median_tpot_ms": med_tpot,
                "mean_tpot_ms": mean_tpot,
                "p99_tpot_ms": p99_tpot,
                "mean_nlat_ms": mean_nlat,
                "median_nlat_ms": med_nlat,
                "date": d.get("date"),
            })
    return pd.DataFrame(rows)


def autosize(ws, df, index: bool = False) -> None:
    offset = 0
    if index:
        idx_w = max([len(str(df.index.name or ""))]
                    + [len(str(v)) for v in df.index]) + 2
        ws.column_dimensions[get_column_letter(1)].width = min(idx_w, 40)
        offset = 1
    for i, col in enumerate(df.columns, start=1 + offset):
        w = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(w, 40)


def write_compare(writer, df: pd.DataFrame) -> None:
    """nice_compare: per-rate ngang dual vs opt với ratio + winner."""
    metrics = [
        ("median_ttft_ms", True),
        ("mean_ttft_ms", True),
        ("p99_ttft_ms", True),
        ("median_tpot_ms", True),
        ("mean_tpot_ms", True),
        ("p99_tpot_ms", True),
        ("mean_nlat_ms", True),
        ("median_nlat_ms", True),
        ("output_throughput", False),  # higher = better
        ("request_throughput", False),
    ]
    rows = []
    for r in RATES:
        for m, lower_better in metrics:
            dual = df[(df.scheduler == "dual1.0") & (df.qps == r)][m]
            opt = df[(df.scheduler == "opt-xxx") & (df.qps == r)][m]
            d_v = float(dual.iloc[0]) if len(dual) else float("nan")
            o_v = float(opt.iloc[0]) if len(opt) else float("nan")
            ratio = d_v / o_v if o_v else float("nan")
            if np.isnan(ratio):
                winner = "—"
            elif lower_better:
                winner = "DPS" if ratio < 1 else ("opt-xxx" if ratio > 1 else "tie")
            else:
                winner = "DPS" if ratio > 1 else ("opt-xxx" if ratio < 1 else "tie")
            gap_pct = (ratio - 1) * 100 if not np.isnan(ratio) else float("nan")
            rows.append({
                "rate": int(r),
                "metric": m,
                "lower_is_better": lower_better,
                "DPS (dual1.0)": round(d_v, 2) if not np.isnan(d_v) else "OOM",
                "opt-xxx (Fu et al)": round(o_v, 2) if not np.isnan(o_v) else "OOM",
                "ratio d/o": round(ratio, 4) if not np.isnan(ratio) else "—",
                "gap %": round(gap_pct, 1) if not np.isnan(gap_pct) else "—",
                "winner": winner,
            })
    nice = pd.DataFrame(rows)
    nice.to_excel(writer, sheet_name="nice_compare", index=False)

    ws = writer.sheets["nice_compare"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    dps_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    opt_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    tie_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    light_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    ncols = len(nice.columns)
    n_per_rate = len(metrics)
    for ci in range(1, ncols + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for ri in range(2, len(rows) + 2):
        rate_idx = (ri - 2) // n_per_rate
        bg = alt_fill if rate_idx % 2 == 0 else light_fill
        winner = rows[ri - 2]["winner"]
        for ci in range(1, ncols + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = bg
            c.border = thin
            c.alignment = Alignment(horizontal="center" if ci <= 3 else "right",
                                     vertical="center")
        # color winner cell
        wcell = ws.cell(row=ri, column=ncols)
        if winner == "DPS":
            wcell.fill = dps_fill
            wcell.font = Font(bold=True, color="006100")
        elif winner == "opt-xxx":
            wcell.fill = opt_fill
            wcell.font = Font(bold=True, color="9C0006")
        else:
            wcell.fill = tie_fill

    autosize(ws, nice)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default=".",
                    help="Folder chứa SERVE_DUAL_TEST/ và SERVE_OPTXXX/")
    ap.add_argument("--out", required=True, help="Output xlsx")
    args = ap.parse_args()

    base = Path(args.base)
    out = Path(args.out)

    df = load_all(base)
    print(f"Loaded {len(df)} runs")
    print(df[["scheduler", "qps", "duration", "completed",
              "median_tpot_ms", "mean_nlat_ms"]].to_string(index=False))

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        all_runs_cols = (["scheduler", "qps", "source_dir", "model", "num_prompts"]
                         + METRIC_COLS + ["date", "file"])
        df_sorted = df.sort_values(["scheduler", "qps"])
        df_sorted[all_runs_cols].to_excel(writer, sheet_name="all_runs", index=False)
        autosize(writer.sheets["all_runs"], df_sorted[all_runs_cols])

        write_compare(writer, df)

        qps_order = sorted(df["qps"].unique())
        sched_order = ["dual1.0", "opt-xxx"]
        for metric in PIVOT_SHEETS:
            piv = (df.pivot_table(index="qps", columns="scheduler",
                                  values=metric, aggfunc="mean")
                   .reindex(index=qps_order, columns=sched_order))
            piv["ratio d/o"] = piv["dual1.0"] / piv["opt-xxx"]
            piv.to_excel(writer, sheet_name=metric)
            autosize(writer.sheets[metric], piv, index=True)

    print(f"\nWrote {out}")


if __name__ == "__main__":
    main()
