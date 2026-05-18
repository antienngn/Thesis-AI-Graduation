"""build_ser_res_async_merge_xlsx.py — generate ser_res_async_merge.xlsx.

Tạo file Excel có cấu trúc:
  - all_runs               : long format, 1 row per (scheduler, qps)
  - p99_ttft_ms            : pivot (rows=qps, cols=scheduler)
  - mean_ttft_ms, median_ttft_ms
  - p99_tpot_ms, mean_tpot_ms
  - output_throughput, request_throughput
  - n_latency_ms           : pivot N_latency (= (ttft + sum(itls)) / output_len)
  - nice_format            : ★ NEW sheet — định dạng đẹp giống screenshot,
                             group theo Request/rate, 5 schedulers/group.

Tương tự serve_res.xlsx nhưng exclude opt-cpu-warmup2.0 và tpt-class10-xxx
(per yêu cầu user — chỉ giữ 5 schedulers cho nice_format).
"""
import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Schedulers cho nice_format sheet (theo thứ tự display)
NICE_SCHEDULERS = [
    ("fcfs",                      "First-come-first-serve"),
    ("sjf",                       "Shortest-job-first"),
    ("srtf",                      "Shortest-remaining job first (Oracle)"),
    ("opt-xxx",                   "Ranking scheduler (paper)"),
    ("opt-cpu-async-merged1.0",   "Fcfs + Cpu rank (ours)"),
]
OURS_KEY = "opt-cpu-async-merged1.0"

# All schedulers cho pivot sheets (giữ tpt-class10-xxx nếu có)
ALL_SCHEDULER_ORDER = [
    "fcfs", "sjf", "srtf", "tpt-class10-xxx", "opt-xxx",
    "opt-cpu-async-merged1.0",
]

# Cột metrics cho all_runs + pivot sheets (match serve_res.xlsx)
METRIC_COLS = [
    "duration", "completed", "request_throughput",
    "input_throughput", "output_throughput",
    "mean_ttft_ms", "median_ttft_ms", "p99_ttft_ms",
    "mean_tpot_ms", "median_tpot_ms", "p99_tpot_ms",
    "n_latency_ms",
]

# Pivot sheets (key = JSON field, sheet_name)
PIVOT_SHEETS = [
    "p99_ttft_ms", "mean_ttft_ms", "median_ttft_ms",
    "p99_tpot_ms", "mean_tpot_ms",
    "output_throughput", "request_throughput",
    "n_latency_ms",
]


def compute_n_latency_ms(data: dict) -> float:
    """N_latency = mean( (ttft + sum(itl)) / output_len ) * 1000 (ms).

    Match aggregate_nlatency.py formula (chính xác hơn benchmark_serving_real
    vì không bao gồm [DONE] chunk overhead).
    """
    ttfts = data.get("ttfts") or []
    itls = data.get("itls") or []
    output_lens = data.get("output_lens") or []
    if not (ttfts and itls and output_lens):
        return float("nan")
    nlat = []
    for i in range(len(ttfts)):
        if output_lens[i] > 0:
            latency = ttfts[i] + sum(itls[i])
            nlat.append(latency / output_lens[i])
    if not nlat:
        return float("nan")
    return float(np.mean(nlat) * 1000.0)


def load_results(src: Path) -> pd.DataFrame:
    """Load tất cả vllm-*.json → DataFrame long format."""
    files = sorted(src.glob("vllm-*.json"))
    if not files:
        raise SystemExit(f"No vllm-*.json found in {src}")
    rows = []
    for f in files:
        with f.open() as fh:
            d = json.load(fh)
        row = {
            "scheduler": d.get("schedule_type"),
            "qps": float(d.get("request_rate")),
            "cv": float(d.get("cv", 1.0)) if d.get("cv") is not None else 1.0,
            "model": d.get("model_id"),
            "num_prompts": d.get("num_prompts"),
            "completed": d.get("completed"),
            "duration": d.get("duration"),
            "request_throughput": d.get("request_throughput"),
            "input_throughput": d.get("input_throughput"),
            "output_throughput": d.get("output_throughput"),
            "mean_ttft_ms": d.get("mean_ttft_ms"),
            "median_ttft_ms": d.get("median_ttft_ms"),
            "p99_ttft_ms": d.get("p99_ttft_ms"),
            "mean_tpot_ms": d.get("mean_tpot_ms"),
            "median_tpot_ms": d.get("median_tpot_ms"),
            "p99_tpot_ms": d.get("p99_tpot_ms"),
            "n_latency_ms": compute_n_latency_ms(d),
            "date": d.get("date"),
            "file": f.name,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def order_schedulers(scheds: list) -> list:
    """Sort theo ALL_SCHEDULER_ORDER trước, rồi extra."""
    known = [s for s in ALL_SCHEDULER_ORDER if s in scheds]
    extra = sorted(s for s in scheds if s not in ALL_SCHEDULER_ORDER)
    return known + extra


def autosize(ws, df, index: bool = False) -> None:
    """Auto-size columns based on content."""
    offset = 0
    if index:
        idx_width = max([len(str(df.index.name or ""))] +
                        [len(str(v)) for v in df.index]) + 2
        ws.column_dimensions[get_column_letter(1)].width = min(idx_width, 40)
        offset = 1
    for i, col in enumerate(df.columns, start=1 + offset):
        width = max([len(str(col))] +
                    [len(str(v)) for v in df[col].astype(str)]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)


def write_nice_format(writer, df: pd.DataFrame) -> None:
    """Write 'nice_format' sheet — table giống screenshot.

    Layout:
      Request/rate | Scheduler                           | metrics...
      2            | First-come-first-serve              | xxx...
      2            | Shortest-job-first                  | xxx...
      2            | Shortest-remaining job first (Oracle)| xxx...
      2            | Ranking scheduler (paper)            | xxx...
      2            | Fcfs + Cpu rank (ours)              | xxx...   ← BOLD
      4            | ...
      ...

    Style:
      - Bold header
      - Bold "ours" row
      - Alternating background per rate group
      - Center-align numbers
      - Border on group boundaries
    """
    cols = ["Request/rate", "Scheduler", "duration (second)", "N_latency",
            "mean_ttft_ms", "median_ttft_ms", "p99_ttft_ms",
            "mean_tpot_ms", "median_tpot_ms", "p99_tpot_ms"]

    # Build long table rows
    rates = sorted(df["qps"].unique())
    rows_data = []
    for rate in rates:
        for sched_key, sched_display in NICE_SCHEDULERS:
            mask = (df["scheduler"] == sched_key) & (df["qps"] == rate)
            sub = df[mask]
            if sub.empty:
                row = [rate, sched_display, "OOM", "OOM", "OOM", "OOM",
                       "OOM", "OOM", "OOM", "OOM"]
            else:
                r = sub.iloc[0]
                row = [
                    rate, sched_display,
                    round(r["duration"], 2)
                        if pd.notna(r["duration"]) else "OOM",
                    round(r["n_latency_ms"], 2)
                        if pd.notna(r["n_latency_ms"]) else "OOM",
                    round(r["mean_ttft_ms"], 2)
                        if pd.notna(r["mean_ttft_ms"]) else "OOM",
                    round(r["median_ttft_ms"], 2)
                        if pd.notna(r["median_ttft_ms"]) else "OOM",
                    round(r["p99_ttft_ms"], 2)
                        if pd.notna(r["p99_ttft_ms"]) else "OOM",
                    round(r["mean_tpot_ms"], 2)
                        if pd.notna(r["mean_tpot_ms"]) else "OOM",
                    round(r["median_tpot_ms"], 2)
                        if pd.notna(r["median_tpot_ms"]) else "OOM",
                    round(r["p99_tpot_ms"], 2)
                        if pd.notna(r["p99_tpot_ms"]) else "OOM",
                ]
            rows_data.append(row)

    nice_df = pd.DataFrame(rows_data, columns=cols)
    nice_df.to_excel(writer, sheet_name="nice_format", index=False)

    # === Apply styling via openpyxl ===
    ws = writer.sheets["nice_format"]

    # Header style
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data rows — alternating fill per rate group + bold "ours" + borders
    n_per_group = len(NICE_SCHEDULERS)
    light_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    alt_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    ours_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    oom_font = Font(color="DC143C", italic=True)

    for row_idx in range(2, len(rows_data) + 2):
        group_idx = (row_idx - 2) // n_per_group  # 0-based group
        row_in_group = (row_idx - 2) % n_per_group
        is_ours = row_in_group == n_per_group - 1  # last row of each group

        # Background fill: ours = yellow, others alt by group
        fill = ours_fill if is_ours else (
            alt_fill if group_idx % 2 == 0 else light_fill
        )

        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            if is_ours:
                cell.font = Font(bold=True)
            # Center numeric columns
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right",
                                           vertical="center")
                # OOM cells get red italic
                if cell.value == "OOM":
                    cell.font = Font(color="DC143C", italic=True,
                                     bold=is_ours)
            else:
                cell.alignment = Alignment(horizontal="center"
                                           if col_idx == 1 else "left",
                                           vertical="center")

    # Auto-size columns
    autosize(ws, nice_df)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default="SER_RES_ASYNC_MERGE",
                    help="Folder chứa vllm-*.json")
    ap.add_argument("--out", default="ser_res_async_merge.xlsx",
                    help="Output xlsx")
    args = ap.parse_args()

    src = Path(args.src)
    out = Path(args.out)

    # Load all data
    df = load_results(src)
    print(f"Loaded {len(df)} runs from {src}")
    print(f"Schedulers: {sorted(df['scheduler'].unique())}")
    print(f"Rates: {sorted(df['qps'].unique())}")

    # Sort by (scheduler order, qps)
    sched_order = order_schedulers(df["scheduler"].unique())
    df["_sched_idx"] = df["scheduler"].map(
        {s: i for i, s in enumerate(sched_order)}
    )
    df = df.sort_values(["_sched_idx", "qps"]).drop(columns=["_sched_idx"])

    qps_order = sorted(df["qps"].unique())

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # === Sheet 1: all_runs (long format) ===
        all_runs_cols = (["scheduler", "qps", "cv", "model", "num_prompts"]
                         + METRIC_COLS + ["date", "file"])
        df[all_runs_cols].to_excel(
            writer, sheet_name="all_runs", index=False
        )
        autosize(writer.sheets["all_runs"], df[all_runs_cols])

        # === Pivot sheets ===
        for metric in PIVOT_SHEETS:
            pivot = (
                df.pivot_table(index="qps", columns="scheduler",
                               values=metric, aggfunc="mean")
                .reindex(index=qps_order, columns=sched_order)
            )
            pivot.to_excel(writer, sheet_name=metric)
            autosize(writer.sheets[metric], pivot, index=True)

        # === Nice format sheet ===
        write_nice_format(writer, df)

    print(f"\nWrote {out}")
    print(f"Sheets: all_runs + {len(PIVOT_SHEETS)} pivots + nice_format")


if __name__ == "__main__":
    main()
