#!/usr/bin/env python3
"""Aggregate kết quả của N runs benchmark_dual.sh → mean ± std per (rate, metric).

Layout input: SRC/run{1,2,...,N}/vllm-*.json
Output: 1 xlsx với nhiều sheet:
    - per_run:  long table, mỗi row = 1 (run, rate, metric set)
    - mean:     pivot index=rate, columns=metric, value=mean across runs
    - std:      pivot index=rate, columns=metric, value=std across runs
    - mean_pm_std:  cell = "mean ± std" để paste vào báo cáo
"""
import argparse
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd

METRICS = [
    "request_throughput",
    "output_throughput",
    "mean_ttft_ms",
    "median_ttft_ms",
    "p99_ttft_ms",
    "mean_tpot_ms",
    "median_tpot_ms",
    "p99_tpot_ms",
    "total_input_tokens",
    "total_output_tokens",
    "completed",
    "duration",
]


def compute_nlatency(d):
    ttfts = d.get("ttfts") or []
    itls = d.get("itls") or []
    olens = d.get("output_lens") or []
    if not (ttfts and itls and olens):
        return None, None
    vals = []
    for i in range(len(ttfts)):
        if olens[i] > 0:
            vals.append((ttfts[i] + sum(itls[i])) / olens[i])
    if not vals:
        return None, None
    arr = np.asarray(vals) * 1000.0  # → ms/token
    return float(arr.mean()), float(np.percentile(arr, 90))


def load_run_dir(run_dir: Path, run_id: int) -> list[dict]:
    rows = []
    for f in sorted(run_dir.glob("vllm-*.json")):
        with f.open() as fh:
            d = json.load(fh)
        nlat_mean, nlat_p90 = compute_nlatency(d)
        row = {
            "run": run_id,
            "qps": float(d.get("request_rate")),
            "scheduler": d.get("schedule_type"),
            "mean_nlatency_ms": nlat_mean,
            "p90_nlatency_ms": nlat_p90,
            "_file": f.name,
        }
        for m in METRICS:
            row[m] = d.get(m)
        rows.append(row)
    return rows


def autosize(ws, df, index=False):
    from openpyxl.utils import get_column_letter
    offset = 0
    if index:
        idx_w = max([len(str(df.index.name or ""))] +
                    [len(str(v)) for v in df.index]) + 2
        ws.column_dimensions[get_column_letter(1)].width = min(idx_w, 40)
        offset = 1
    for i, col in enumerate(df.columns, start=1 + offset):
        w = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="dir chứa run{i}/ subdirs")
    ap.add_argument("--out", required=True, help="output xlsx")
    args = ap.parse_args()

    src = Path(args.src)
    run_dirs = sorted([p for p in src.iterdir() if p.is_dir()
                       and re.match(r"run\d+$", p.name)])
    if not run_dirs:
        raise SystemExit(f"No run*/ subdirs in {src}")

    all_rows = []
    for rd in run_dirs:
        rid = int(rd.name[3:])
        rows = load_run_dir(rd, rid)
        if not rows:
            print(f"  [WARN] {rd}: no vllm-*.json")
            continue
        all_rows.extend(rows)
        print(f"  {rd.name}: {len(rows)} files, qps={sorted({r['qps'] for r in rows})}")

    df = pd.DataFrame(all_rows)
    if df.empty:
        raise SystemExit("No data")

    print(f"\nTotal: {len(df)} rows, {df['run'].nunique()} runs, "
          f"qps={sorted(df['qps'].unique())}, "
          f"sched={list(df['scheduler'].unique())}")

    metric_cols = ["mean_nlatency_ms", "p90_nlatency_ms"] + METRICS
    metric_cols = [c for c in metric_cols if c in df.columns]

    mean_df = (df.groupby("qps")[metric_cols]
                 .mean()
                 .sort_index())
    std_df = (df.groupby("qps")[metric_cols]
                .std(ddof=1)
                .sort_index())

    # mean ± std as string for human-readable sheet
    def fmt(m, s):
        if pd.isna(m):
            return ""
        if pd.isna(s):
            return f"{m:.2f}"
        return f"{m:.2f} ± {s:.2f}"

    pm_df = pd.DataFrame(
        {c: [fmt(mean_df.loc[r, c], std_df.loc[r, c]) for r in mean_df.index]
         for c in mean_df.columns},
        index=mean_df.index,
    )

    out = Path(args.out)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="per_run", index=False)
        autosize(writer.sheets["per_run"], df)

        mean_df.to_excel(writer, sheet_name="mean")
        autosize(writer.sheets["mean"], mean_df, index=True)

        std_df.to_excel(writer, sheet_name="std")
        autosize(writer.sheets["std"], std_df, index=True)

        pm_df.to_excel(writer, sheet_name="mean_pm_std")
        autosize(writer.sheets["mean_pm_std"], pm_df, index=True)

    print(f"\nWrote {out}")
    print("\n=== mean (key metrics) ===")
    key = [c for c in ["mean_nlatency_ms", "p90_nlatency_ms",
                       "mean_ttft_ms", "mean_tpot_ms",
                       "request_throughput", "output_throughput"]
           if c in mean_df.columns]
    print(mean_df[key].round(2).to_string())


if __name__ == "__main__":
    main()
