#!/usr/bin/env python3
"""build_predictor_latency_xlsx.py — Đọc predictor_latency_raw.csv,
build xlsx 4 sheet để paste vào slide.

Sheets:
  - all_calls:           raw, 1 row = 1 forward call
  - mean_latency_ms:     pivot row=rate, col=(sched/predictor)
  - median_latency_ms:   pivot row=rate, col=(sched/predictor)
  - p99_latency_ms:      pivot row=rate, col=(sched/predictor)
  - nice_format:         bảng tổng hợp + cross-ref TTFT

Cross-ref TTFT: lấy mean_ttft_ms từ vllm-*.json để tính
% TTFT = (per_req_ms_predictor / mean_ttft_ms) * 100
"""

import csv
import json
from collections import defaultdict
from glob import glob
from pathlib import Path

import numpy as np
import openpyxl

import os as _os

ROOT = Path(__file__).parent
# Override qua env LOG_DIR để build cho thư mục khác (vd TEMP_RES_ASYNC_BURST):
#   LOG_DIR=TEMP_RES_ASYNC_BURST python build_predictor_latency_xlsx.py
LOG_DIR = ROOT / _os.environ.get("LOG_DIR", "PRE_LAT_E2E")
IN_CSV = LOG_DIR / "predictor_latency_raw.csv"
OUT_XLSX = LOG_DIR / "predictor_latency_e2e.xlsx"


def main():
    if not IN_CSV.exists():
        print(f"Missing {IN_CSV}. Chạy parse_predictor_latency.py trước.")
        return

    with open(IN_CSV) as f:
        raw = list(csv.DictReader(f))
    if not raw:
        print(f"{IN_CSV} empty.")
        return

    # ----- Filter: chỉ giữ OPT-125m (GPU AUXLLM via opt-xxx + CPU OV via
    # async-merged). Bỏ Pythia khỏi tổng hợp theo yêu cầu user.
    raw = [r for r in raw if r["predictor"] == "opt"]
    if not raw:
        print("No 'opt' predictor rows. Check raw CSV.")
        return

    # ----- Group latencies by (sched, predictor, rate) -----
    lat_groups = defaultdict(list)   # key -> [latency_ms]
    n_groups = defaultdict(list)     # key -> [n_per_chunk]
    per_req_groups = defaultdict(list)  # key -> [per_req_ms]
    for r in raw:
        key = (r["scheduler"], r["predictor"], int(r["rate"]))
        lat_groups[key].append(float(r["latency_ms"]))
        n_groups[key].append(int(r["n_per_chunk"]))
        per_req_groups[key].append(float(r["per_req_ms"]))

    # ----- Cross-ref TTFT từ JSON benchmark -----
    # Một (sched, rate) có thể có nhiều JSON từ các predictor khác nhau (cùng
    # schedule_type) nhưng KHÔNG phân biệt được trong JSON. Match qua server
    # log mtime: với mỗi (sched, predictor, rate), lấy mtime của server log
    # tương ứng, tìm JSON nào có mtime gần nhất → đó là kết quả bench của
    # predictor đó.
    import os
    # Build candidate JSONs by (sched_key, rate) -> [(mtime, ttft, path)]
    candidates = defaultdict(list)
    for jp in glob(str(LOG_DIR / "vllm-*.json")):
        try:
            with open(jp) as f:
                j = json.load(f)
            sched = j["schedule_type"]
            rate = int(j["request_rate"])
            if sched.startswith("opt-cpu-async-merged"):
                sched_key = "async-merged"
            elif sched == "opt-xxx":
                sched_key = "opt-xxx"
            else:
                sched_key = sched
            candidates[(sched_key, rate)].append(
                (os.path.getmtime(jp), j["mean_ttft_ms"], jp))
        except Exception as e:
            print(f"WARN: cannot parse {jp}: {e}")

    # For each (sched, predictor, rate), find server log mtime and match
    # closest JSON (within group of same sched+rate).
    ttft = {}  # (sched, predictor, rate) -> mean_ttft_ms
    for key in lat_groups:
        sched, predictor, rate = key
        log_path = LOG_DIR / f"server_{sched}_{predictor}_r{rate}.log"
        if not log_path.exists():
            continue
        log_mtime = os.path.getmtime(log_path)
        cands = candidates.get((sched, rate), [])
        if not cands:
            continue
        # Pick JSON with smallest |mtime - log_mtime| difference
        best = min(cands, key=lambda x: abs(x[0] - log_mtime))
        ttft[key] = best[1]

    # ===== Build workbook =====
    wb = openpyxl.Workbook()

    # ---- Sheet 1: all_calls ----
    ws = wb.active
    ws.title = "all_calls"
    ws.append(["scheduler", "predictor", "rate",
               "n_per_chunk", "latency_ms", "per_req_ms", "source"])
    for r in raw:
        ws.append([
            r["scheduler"], r["predictor"], int(r["rate"]),
            int(r["n_per_chunk"]), float(r["latency_ms"]),
            float(r["per_req_ms"]), r.get("source", ""),
        ])

    # ---- Sheet 2-4: pivot tables ----
    cols = sorted({(s, p) for (s, p, _) in lat_groups.keys()})
    rates = sorted({rt for (_, _, rt) in lat_groups.keys()})

    def stat_pivot(name, fn):
        ws = wb.create_sheet(name)
        ws.append(["rate"] + [f"{s}/{p}" for s, p in cols])
        for rate in rates:
            row = [rate]
            for s, p in cols:
                xs = lat_groups.get((s, p, rate), [])
                row.append(fn(xs) if xs else None)
            ws.append(row)

    stat_pivot("mean_latency_ms", np.mean)
    stat_pivot("median_latency_ms", np.median)
    stat_pivot("p99_latency_ms", lambda xs: float(np.percentile(xs, 99)))

    # ---- Sheet 5: nice_format ----
    ws = wb.create_sheet("nice_format")
    ws.append([
        "Request/rate", "Scheduler", "Predictor",
        "# forward calls", "Mean lat (ms)", "Median (ms)",
        "P99 (ms)", "Mean per-req (ms)",
        "Mean TTFT (ms)", "% TTFT contribution",
    ])
    for (s, p, rate) in sorted(lat_groups.keys(),
                                key=lambda k: (k[2], k[0], k[1])):
        xs = lat_groups[(s, p, rate)]
        per_req = per_req_groups[(s, p, rate)]
        mean_lat = float(np.mean(xs))
        median_lat = float(np.median(xs))
        p99_lat = float(np.percentile(xs, 99))
        mean_per_req = float(np.mean(per_req))
        ttft_ms = ttft.get((s, p, rate))
        pct = (mean_per_req / ttft_ms * 100) if ttft_ms else None
        ws.append([
            rate, s, p, len(xs),
            round(mean_lat, 3), round(median_lat, 3), round(p99_lat, 3),
            round(mean_per_req, 3),
            round(ttft_ms, 2) if ttft_ms else None,
            round(pct, 4) if pct is not None else None,
        ])

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"  sheets: all_calls, mean_latency_ms, median_latency_ms, "
          f"p99_latency_ms, nice_format")


if __name__ == "__main__":
    main()
