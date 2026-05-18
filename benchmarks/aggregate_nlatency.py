#!/usr/bin/env python3
"""Aggregate per-request Nlatency from SERVE_RES into a single .xlsx workbook.

Nlatency_i = (ttft_i + sum(itl_i)) / output_len_i        # s/token
Output is in ms/token. Rows = scheduler, columns = request_rate.

Usage:
    python aggregate_nlatency.py [--src SERVE_RES] [--out nlatency.xlsx]
"""
import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd

SCHEDULER_ORDER = ["fcfs", "sjf", "srtf", "tpt-class10-xxx", "opt-xxx", "opt-cpu-warmup2.0"]


def compute_nlatency_ms(data: dict) -> tuple[float | None, float | None, int]:
    ttfts = data.get("ttfts") or []
    itls = data.get("itls") or []
    output_lens = data.get("output_lens") or []
    if not (ttfts and itls and output_lens):
        return None, None, 0
    nlat = []
    for i in range(len(ttfts)):
        if output_lens[i] > 0:
            latency = ttfts[i] + sum(itls[i])
            nlat.append(latency / output_lens[i])
    if not nlat:
        return None, None, 0
    arr = np.asarray(nlat) * 1000.0  # s -> ms
    return float(arr.mean()), float(np.percentile(arr, 90)), len(arr)


def order_schedulers(schedulers: list[str]) -> list[str]:
    known = [s for s in SCHEDULER_ORDER if s in schedulers]
    extra = sorted(s for s in schedulers if s not in SCHEDULER_ORDER)
    return known + extra


def autosize(ws, df, index: bool = False) -> None:
    from openpyxl.utils import get_column_letter

    offset = 0
    if index:
        idx_width = max([len(str(df.index.name or ""))] + [len(str(v)) for v in df.index]) + 2
        ws.column_dimensions[get_column_letter(1)].width = min(idx_width, 40)
        offset = 1
    for i, col in enumerate(df.columns, start=1 + offset):
        width = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(Path(__file__).parent / "SERVE_RES"))
    ap.add_argument("--out", default=str(Path(__file__).parent / "nlatency.xlsx"))
    args = ap.parse_args()

    src = Path(args.src)
    files = sorted(src.glob("vllm-*.json"))
    if not files:
        raise SystemExit(f"No vllm-*.json files found in {src}")

    rows = []
    for p in files:
        with p.open() as f:
            data = json.load(f)
        mean_ms, p90_ms, n = compute_nlatency_ms(data)
        if mean_ms is None:
            print(f"  [WARN] skip {p.name}: missing ttfts/itls/output_lens")
            continue
        rows.append({
            "scheduler": data.get("schedule_type"),
            "qps": float(data.get("request_rate")),
            "mean_nlatency_ms": mean_ms,
            "p90_nlatency_ms": p90_ms,
            "num_requests": n,
            "file": p.name,
        })

    df = pd.DataFrame(rows)
    sched_order = order_schedulers(sorted(df["scheduler"].unique()))
    qps_order = sorted(df["qps"].unique())

    pivot_mean = (
        df.pivot_table(index="scheduler", columns="qps", values="mean_nlatency_ms", aggfunc="mean")
        .reindex(index=sched_order, columns=qps_order)
    )
    pivot_p90 = (
        df.pivot_table(index="scheduler", columns="qps", values="p90_nlatency_ms", aggfunc="mean")
        .reindex(index=sched_order, columns=qps_order)
    )

    out = Path(args.out)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        pivot_mean.to_excel(writer, sheet_name="mean_nlatency_ms")
        autosize(writer.sheets["mean_nlatency_ms"], pivot_mean, index=True)
        pivot_p90.to_excel(writer, sheet_name="p90_nlatency_ms")
        autosize(writer.sheets["p90_nlatency_ms"], pivot_p90, index=True)

        long = df.sort_values(
            ["scheduler", "qps"],
            key=lambda s: s.map({v: i for i, v in enumerate(sched_order)}) if s.name == "scheduler" else s,
        ).reset_index(drop=True)
        long.to_excel(writer, sheet_name="long", index=False)
        autosize(writer.sheets["long"], long)

    print(f"Wrote {out}  ({len(rows)} runs)")
    print(f"Schedulers: {sched_order}")
    print(f"QPS columns: {qps_order}")


if __name__ == "__main__":
    main()