#!/usr/bin/env python3
"""Aggregate vllm-*.json files in SERVE_RES into a single .xlsx workbook.

Usage:
    python aggregate_serve_res.py [--src SERVE_RES] [--out serve_res.xlsx]
"""
import argparse
import json
import re
from pathlib import Path

import pandas as pd

METRIC_COLS = [
    "request_throughput",
    "output_throughput",
    "input_throughput",
    "mean_ttft_ms",
    "median_ttft_ms",
    "p99_ttft_ms",
    "mean_tpot_ms",
    "median_tpot_ms",
    "p99_tpot_ms",
]

PIVOT_METRICS = [
    "p99_ttft_ms",
    "mean_ttft_ms",
    "median_ttft_ms",
    "p99_tpot_ms",
    "mean_tpot_ms",
    "output_throughput",
    "request_throughput",
]

CV_RE = re.compile(r"-cv(?P<cv>[0-9.]+)-")

SCHEDULER_ORDER = ["fcfs", "sjf", "srtf", "tpt-class10-xxx", "opt-xxx", "opt-cpu-warmup2.0"]


def parse_file(path: Path) -> dict | None:
    with path.open() as f:
        data = json.load(f)
    cv_match = CV_RE.search(path.name)
    row = {
        "scheduler": data.get("schedule_type"),
        "qps": float(data.get("request_rate")),
        "cv": float(cv_match.group("cv")) if cv_match else None,
        "model": data.get("model_id"),
        "num_prompts": data.get("num_prompts"),
        "completed": data.get("completed"),
        "duration_s": data.get("duration"),
        "total_input_tokens": data.get("total_input_tokens"),
        "total_output_tokens": data.get("total_output_tokens"),
    }
    for k in METRIC_COLS:
        row[k] = data.get(k)
    row["date"] = data.get("date")
    row["file"] = path.name
    return row


def order_schedulers(schedulers: list[str]) -> list[str]:
    known = [s for s in SCHEDULER_ORDER if s in schedulers]
    extra = sorted(s for s in schedulers if s not in SCHEDULER_ORDER)
    return known + extra


def autosize(ws, df, index: bool = False) -> None:
    from openpyxl.utils import get_column_letter

    offset = 0
    if index:
        idx_width = max([len(str(df.index.name or ""))] + [len(str(v)) for v in df.index]) + 2
        ws.column_dimensions[get_column_letter(1)].width = min(idx_width, 40)
        offset = 1
    for i, col in enumerate(df.columns, start=1 + offset):
        width = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(Path(__file__).parent / "SERVE_RES"))
    ap.add_argument("--out", default=str(Path(__file__).parent / "serve_res.xlsx"))
    args = ap.parse_args()

    src = Path(args.src)
    files = sorted(src.glob("vllm-*.json"))
    if not files:
        raise SystemExit(f"No vllm-*.json files found in {src}")

    rows = [r for p in files if (r := parse_file(p)) is not None]
    skipped = len(files) - len(rows)
    df = pd.DataFrame(rows)

    sched_order = order_schedulers(sorted(df["scheduler"].unique()))
    df["scheduler"] = pd.Categorical(df["scheduler"], categories=sched_order, ordered=True)
    df = df.sort_values(["scheduler", "qps"]).reset_index(drop=True)

    out = Path(args.out)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="all_runs", index=False)
        autosize(writer.sheets["all_runs"], df)

        for metric in PIVOT_METRICS:
            pivot = df.pivot_table(
                index="qps", columns="scheduler", values=metric, aggfunc="mean", observed=True
            )
            pivot = pivot.reindex(columns=sched_order).sort_index()
            sheet = metric[:31]
            pivot.to_excel(writer, sheet_name=sheet)
            autosize(writer.sheets[sheet], pivot, index=True)

    print(f"Wrote {out}  ({len(rows)} runs, {skipped} skipped)")
    print(f"Schedulers: {sched_order}")
    print(f"QPS: {sorted(df['qps'].unique())}")


if __name__ == "__main__":
    main()
