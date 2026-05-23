"""build_ser_res_dual_xlsx.py — generate ser_res_dual.xlsx (nice_format).

Tạo file Excel có cấu trúc giống build_ser_res_async_merge_xlsx.py:
  - all_runs               : long format, 1 row per (scheduler, qps)
  - p99_ttft_ms, mean_ttft_ms, median_ttft_ms
  - p99_tpot_ms, mean_tpot_ms
  - output_throughput, request_throughput
  - n_latency_ms           : pivot
  - nice_format            : ★ table đẹp, 5 schedulers/group, "dual1.0" = ours

Khác bản async_merge: OURS = dual1.0 thay vì opt-cpu-async-merged1.0.
"""
import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NICE_SCHEDULERS = [
    ("fcfs",    "First-come-first-serve"),
    ("sjf",     "Shortest-job-first"),
    ("srtf",    "Shortest-remaining job first (Oracle)"),
    ("opt-xxx", "Ranking scheduler (paper)"),
    ("dual1.0", "Dual CPU/GPU router (ours)"),
]
OURS_KEY = "dual1.0"

ALL_SCHEDULER_ORDER = [
    "fcfs", "sjf", "srtf", "tpt-class10-xxx", "opt-xxx", "dual1.0",
]

METRIC_COLS = [
    "duration", "completed", "request_throughput",
    "input_throughput", "output_throughput",
    "mean_ttft_ms", "median_ttft_ms", "p99_ttft_ms",
    "mean_tpot_ms", "median_tpot_ms", "p99_tpot_ms",
    "n_latency_ms",
]

PIVOT_SHEETS = [
    "p99_ttft_ms", "mean_ttft_ms", "median_ttft_ms",
    "p99_tpot_ms", "mean_tpot_ms",
    "output_throughput", "request_throughput",
    "n_latency_ms",
]


def compute_n_latency_ms(data: dict) -> float:
    # Use RAW streamed token count = len(itls[i]) + 1.
    ttfts = data.get("ttfts") or []
    itls = data.get("itls") or []
    if not (ttfts and itls):
        return float("nan")
    nlat = []
    for i in range(len(ttfts)):
        n_raw = len(itls[i]) + 1
        if n_raw > 0:
            latency = ttfts[i] + sum(itls[i])
            nlat.append(latency / n_raw)
    if not nlat:
        return float("nan")
    return float(np.mean(nlat) * 1000.0)


def compute_tpot_ms(data: dict) -> tuple[float, float, float]:
    """Recompute TPOT từ raw: TPOT_i = sum(itls[i]) / (n_raw - 1).

    n_raw = len(itls[i]) + 1, nên (n_raw - 1) = len(itls[i]).
    → TPOT_i = sum(itls[i]) / len(itls[i]) = mean(itls[i]).
    """
    itls = data.get("itls") or []
    if not itls:
        return float("nan"), float("nan"), float("nan")
    tpots = []
    for itl in itls:
        if len(itl) > 0:
            tpots.append(sum(itl) / len(itl))
    if not tpots:
        return float("nan"), float("nan"), float("nan")
    arr = np.asarray(tpots) * 1000.0
    return (float(arr.mean()),
            float(np.median(arr)),
            float(np.percentile(arr, 99)))


def load_results(src: Path) -> pd.DataFrame:
    files = sorted(src.glob("vllm-*.json"))
    if not files:
        raise SystemExit(f"No vllm-*.json found in {src}")
    rows = []
    for f in files:
        with f.open() as fh:
            d = json.load(fh)
        mean_tpot, median_tpot, p99_tpot = compute_tpot_ms(d)
        rows.append({
            "scheduler": d.get("schedule_type"),
            "qps": float(d.get("request_rate")),
            "cv": float(d.get("cv", 1.0)) if d.get("cv") is not None else 1.0,
            "model": d.get("model_id"),
            "num_prompts": d.get("num_prompts"),
            "completed": d.get("completed"),
            "duration": d.get("duration"),
            "request_throughput": d.get("request_throughput"),
            "input_throughput": d.get("input_throughput"),
            # Output throughput dùng raw: tổng raw token / duration
            "output_throughput": (
                sum(len(x) + 1 for x in (d.get("itls") or []))
                / d.get("duration") if d.get("duration") else None
            ),
            "mean_ttft_ms": d.get("mean_ttft_ms"),
            "median_ttft_ms": d.get("median_ttft_ms"),
            "p99_ttft_ms": d.get("p99_ttft_ms"),
            # TPOT recomputed từ raw (KHÔNG dùng re-tok output_lens)
            "mean_tpot_ms": mean_tpot,
            "median_tpot_ms": median_tpot,
            "p99_tpot_ms": p99_tpot,
            "n_latency_ms": compute_n_latency_ms(d),
            "date": d.get("date"),
            "file": f.name,
        })
    return pd.DataFrame(rows)


def order_schedulers(scheds: list) -> list:
    known = [s for s in ALL_SCHEDULER_ORDER if s in scheds]
    extra = sorted(s for s in scheds if s not in ALL_SCHEDULER_ORDER)
    return known + extra


def autosize(ws, df, index: bool = False) -> None:
    offset = 0
    if index:
        idx_width = max([len(str(df.index.name or ""))] +
                        [len(str(v)) for v in df.index]) + 2
        ws.column_dimensions[get_column_letter(1)].width = min(idx_width, 40)
        offset = 1
    for i, col in enumerate(df.columns, start=1 + offset):
        width = max([len(str(col))] +
                    [len(str(v)) for v in df[col].astype(str)]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)


def write_nice_format(writer, df: pd.DataFrame) -> None:
    cols = ["Request/rate", "Scheduler", "duration (second)", "N_latency",
            "mean_ttft_ms", "median_ttft_ms", "p99_ttft_ms",
            "mean_tpot_ms", "median_tpot_ms", "p99_tpot_ms"]

    rates = sorted(df["qps"].unique())
    rows_data = []
    for rate in rates:
        for sched_key, sched_display in NICE_SCHEDULERS:
            mask = (df["scheduler"] == sched_key) & (df["qps"] == rate)
            sub = df[mask]
            if sub.empty:
                row = [rate, sched_display] + ["OOM"] * 8
            else:
                r = sub.iloc[0]
                row = [
                    rate, sched_display,
                    round(r["duration"], 2) if pd.notna(r["duration"]) else "OOM",
                    round(r["n_latency_ms"], 2) if pd.notna(r["n_latency_ms"]) else "OOM",
                    round(r["mean_ttft_ms"], 2) if pd.notna(r["mean_ttft_ms"]) else "OOM",
                    round(r["median_ttft_ms"], 2) if pd.notna(r["median_ttft_ms"]) else "OOM",
                    round(r["p99_ttft_ms"], 2) if pd.notna(r["p99_ttft_ms"]) else "OOM",
                    round(r["mean_tpot_ms"], 2) if pd.notna(r["mean_tpot_ms"]) else "OOM",
                    round(r["median_tpot_ms"], 2) if pd.notna(r["median_tpot_ms"]) else "OOM",
                    round(r["p99_tpot_ms"], 2) if pd.notna(r["p99_tpot_ms"]) else "OOM",
                ]
            rows_data.append(row)

    nice_df = pd.DataFrame(rows_data, columns=cols)
    nice_df.to_excel(writer, sheet_name="nice_format", index=False)

    ws = writer.sheets["nice_format"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    n_per_group = len(NICE_SCHEDULERS)
    light_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ours_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for row_idx in range(2, len(rows_data) + 2):
        group_idx = (row_idx - 2) // n_per_group
        row_in_group = (row_idx - 2) % n_per_group
        is_ours = row_in_group == n_per_group - 1
        fill = ours_fill if is_ours else (alt_fill if group_idx % 2 == 0 else light_fill)
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            if is_ours:
                cell.font = Font(bold=True)
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if cell.value == "OOM":
                    cell.font = Font(color="DC143C", italic=True, bold=is_ours)
            else:
                cell.alignment = Alignment(
                    horizontal="center" if col_idx == 1 else "left",
                    vertical="center",
                )

    autosize(ws, nice_df)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="Folder chứa vllm-*.json")
    ap.add_argument("--out", required=True, help="Output xlsx")
    args = ap.parse_args()

    src = Path(args.src)
    out = Path(args.out)

    df = load_results(src)
    print(f"Loaded {len(df)} runs from {src}")
    print(f"Schedulers: {sorted(df['scheduler'].unique())}")
    print(f"Rates: {sorted(df['qps'].unique())}")

    sched_order = order_schedulers(df["scheduler"].unique())
    df["_sched_idx"] = df["scheduler"].map({s: i for i, s in enumerate(sched_order)})
    df = df.sort_values(["_sched_idx", "qps"]).drop(columns=["_sched_idx"])

    qps_order = sorted(df["qps"].unique())

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        all_runs_cols = (["scheduler", "qps", "cv", "model", "num_prompts"]
                         + METRIC_COLS + ["date", "file"])
        df[all_runs_cols].to_excel(writer, sheet_name="all_runs", index=False)
        autosize(writer.sheets["all_runs"], df[all_runs_cols])

        for metric in PIVOT_SHEETS:
            pivot = (df.pivot_table(index="qps", columns="scheduler",
                                    values=metric, aggfunc="mean")
                     .reindex(index=qps_order, columns=sched_order))
            pivot.to_excel(writer, sheet_name=metric)
            autosize(writer.sheets[metric], pivot, index=True)

        write_nice_format(writer, df)

    print(f"\nWrote {out}")
    print(f"Sheets: all_runs + {len(PIVOT_SHEETS)} pivots + nice_format")


if __name__ == "__main__":
    main()
